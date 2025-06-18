from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import Font
from deep_translator import GoogleTranslator
from tqdm import tqdm

# 加载 Excel 文件
workbook = load_workbook('D:\\工作文件\\18031\FqEN18031\\Template_EN18031_TechnicalDoc.v1.3.xlsx')

# 统计需要翻译的单元格总数
total_cells = 0
for sheet in workbook.worksheets:
    for row in sheet.iter_rows():
        for _ in row:
            total_cells += 1

# 使用 tqdm 显示进度条
with tqdm(total=total_cells, desc='翻译进度') as pbar:
    # 遍历每个工作表
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                # 检查单元格是否有值并且值为字符串
                if cell.value and isinstance(cell.value, str):
                    try:
                        # 使用 GoogleTranslator 进行翻译，源语言自动检测，目标语言为中文
                        translated_text = GoogleTranslator(source='auto', target='zh - CN').translate(cell.value)
                        cell.value = translated_text
                        # 设置字体为宋体
                        cell.font = Font(name='宋体')
                    except Exception as e:
                        print(f"翻译单元格 {cell.coordinate} 时出错: {e}")
                # 更新进度条
                pbar.update(1)

# 保存修改后的 Excel 文件
translated_file_path = 'C:\\Users\\cvter\\Desktop\\Template_EN18031_TechnicalDoc.v1.3_translated.xlsx'
workbook.save(translated_file_path)
